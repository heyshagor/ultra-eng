#!/usr/bin/env python3
"""
Test script for PDF Data Extractor
Run this to verify installation and see demo output
"""

import sys
import os
from pathlib import Path

# Add package to path
sys.path.insert(0, str(Path(__file__).parent))

def test_imports():
    """Test all imports"""
    print("[TEST 1/5] Testing imports...")
    try:
        from pdf_data_extractor.extractor import PDFExtractor, extract_pdf_to_excel
        from pdf_data_extractor.parser import BTMAParser
        import pandas as pd
        print("  ✓ All imports successful")
        return True
    except ImportError as e:
        print(f"  ✗ Import error: {e}")
        return False

def test_sample_data():
    """Create and test sample data"""
    print("\n[TEST 2/5] Creating sample data...")
    try:
        import pandas as pd
        from test_extractor import create_demo_data

        df = create_demo_data()
        print(f"  ✓ Sample data created: {len(df)} records")

        # Save demo files
        df.to_excel('demo_sample.xlsx', index=False)
        df.to_csv('demo_sample.csv', index=False, encoding='utf-8-sig')
        print("  ✓ Demo files: demo_sample.xlsx, demo_sample.csv")

        return df
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return None

def test_parser():
    """Test parser functionality"""
    print("\n[TEST 3/5] Testing parser...")
    try:
        from pdf_data_extractor.parser import BTMAParser
        import pandas as pd

        parser = BTMAParser()

        # Create sample data for parser
        sample_text = """
        ABM Industries Ltd.
        House # 06, Road # 23, Gulshan-1, Dhaka-1212
        Phone: +88 02 9882368
        Email: info@abmind.com
        Website: www.abmind.com
        Products: Poplin, Twill, Oxford, Canvas

        Apex Textile Mills Ltd.
        185, Nayabazsr, Narayanganj-1400
        Phone: +88 02 7646215
        Email: contact@apex-textile.com
        Products: Single Jersey, Double Jersey
        """

        # Simulate extractor data
        class MockExtractor:
            def __init__(self):
                # Provide a minimal pages_data structure
                self.pages_data = [{
                    'page_number': 1,
                    'text': sample_text,
                    'tables': [],
                    'dimensions': {'width': 0, 'height': 0}
                }]

            def _structure_output(self):
                return {
                    'full_text': sample_text,
                    'metadata': {'total_pages': 1, 'file_name': 'test.pdf'},
                    'pages': self.pages_data
                }

        parser.load_from_extractor(MockExtractor())
        parser.parse_all()

        print(f"  ✓ Parser working:")
        for name, df in parser.parsed_data.items():
            print(f"    - {name}: {len(df)} records")

        return True
    except Exception as e:
        print(f"  ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        return False

def test_excel_export():
    """Test Excel export"""
    print("\n[TEST 4/5] Testing Excel export...")
    try:
        import pandas as pd
        from openpyxl import load_workbook

        # Create test data
        df = pd.DataFrame({
            'Company': ['Test Co', 'Demo Ltd'],
            'Phone': ['+88012345678', '+88098765432'],
            'Email': ['test@example.com', 'demo@example.com']
        })

        # Test write
        output = 'test_output.xlsx'
        df.to_excel(output, index=False, engine='openpyxl')

        # Test read
        wb = load_workbook(output)
        ws = wb.active
        count = ws.max_row

        print(f"  ✓ Excel I/O working: {count} rows")
        os.remove(output)

        return True
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return False

def test_cli():
    """Test CLI execution"""
    print("\n[TEST 5/5] Testing CLI...")
    try:
        import subprocess

        result = subprocess.run(
            [sys.executable, 'main.py', '--test'],
            capture_output=True,
            text=True,
            timeout=30
        )

        if result.returncode == 0:
            print("  ✓ CLI working")
            print("  ✓ Sample data generated")
            # Clean up test files
            for f in ['btma_sample_parsed.xlsx', 'btma_fabric_data.xlsx', 'btma_fabric_data.csv']:
                if os.path.exists(f):
                    os.remove(f)
            return True
        else:
            print(f"  ✗ CLI error: {result.stderr}")
            return False
    except Exception as e:
        print(f"  ✗ Error: {e}")
        return False

def create_demo_data():
    """Create demo BTMA data"""
    import pandas as pd

    data = {
        'Manufacturer Name': [
            'ABM Industries Ltd.',
            'Apex Textile Mills Ltd.',
            'Beximco Textiles Ltd.',
            'Square Textiles Ltd.',
            'Pacific Textiles Ltd.',
            'Fakir Group',
            'Noman Group',
            'M.M. Textiles Ltd.',
            'H.R.Textiles',
            'DBL Group'
        ],
        'Address': [
            'House # 06, Road # 23, Gulshan-1, Dhaka-1212',
            '185, Nayabazsr, Narayanganj-1400',
            'Beximco Industrial Park, Dhaka-Barisal Road, Dhaka',
            'Square Centre, 274, Faiz Ahmed Faiz Road, Panthapath, Dhaka',
            'House # 08, Road # 02, Banani, Dhaka-1213',
            'Fakir Chamber, 49, Dilkusha C/A, Motijheel, Dhaka-1000',
            'Noman Centre, 273,/D, Bir Uttam CR Dutta Road, Dhaka-1205',
            'M.M. Tower, 28, Kemal Ataturk Avenue, Banani, Dhaka',
            'H.R. Tower, 30, Shaymoli, Dhaka',
            'DBL Centre, 27, Bijoy Nagar, Dhaka-1000'
        ],
        'Phone': [
            '+88 02 9882368',
            '+88 02 7646215',
            '+88 02 9005025',
            '+88 02 9120124',
            '+88 02 8851041',
            '+88 02 9562241',
            '+88 02 8612461',
            '+88 02 8953251',
            '+88 02 9141022',
            '+88 02 9123456'
        ],
        'Email': [
            'info@abmind.com',
            'contact@apex-textile.com',
            'info@beximco.com',
            'info@squaretextiles.com',
            'info@pacific-textiles.com',
            'info@fakirgroup.com',
            'contact@nomangroup.com',
            'info@mmtextiles.com',
            'info@hrtextiles.com',
            'info@dblgroup.com'
        ],
        'Website': [
            'www.abmind.com',
            'www.apex-textile.com',
            'www.beximco.com',
            'www.squaretextiles.com',
            'www.pacific-textiles.com',
            'www.fakirgroup.com',
            'www.nomangroup.com',
            'www.mmtextiles.com',
            'www.hrtextiles.com',
            'www.dblgroup.com'
        ],
        'Product Category': [
            'Woven Fabrics',
            'Knitted Fabrics',
            'Woven & Knitted',
            'Denim Fabrics',
            'Woven Fabrics',
            'Textile Processing',
            'Complete Textile Solution',
            'Fabrics & Garments',
            'Knit Fabrics',
            'Textile & Garments'
        ],
        'Key Products': [
            'Poplin, Twill, Oxford, Canvas',
            'Single Jersey, Double Jersey, Interlock',
            'Woven, Knit, Denim, Terry',
            'Denim, Comforters, Shirting',
            'Plain Weave, Twill, Satin',
            'Processing, Weaving, Knitting',
            'Spinning, Weaving, Processing, Garments',
            'Woven Shirtings, Suitings',
            'Pique, Rib, Lacoste',
            'Knit, Woven, Sweater, Garments'
        ],
    }

    return pd.DataFrame(data)

def main():
    print("="*60)
    print(" BTMA PDF Data Extractor - Test Suite")
    print("="*60)

    results = []

    # Run tests
    results.append(("Imports", test_imports()))
    results.append(("Sample Data", test_sample_data() is not None))
    results.append(("Parser", test_parser()))
    results.append(("Excel Export", test_excel_export()))
    results.append(("CLI", test_cli()))

    # Summary
    print("\n" + "="*60)
    print(" TEST SUMMARY")
    print("="*60)

    passed = sum(1 for _, r in results if r)
    total = len(results)

    for test_name, result in results:
        status = "✓ PASS" if result else "✗ FAIL"
        print(f"{test_name:.<40} {status}")

    print("="*60)
    print(f"Result: {passed}/{total} tests passed")
    print("="*60)

    if passed == total:
        print("\n All tests passed! The package is ready to use.\n")
        print("Next steps:")
        print("1. Place your PDF in the current directory")
        print("2. Run: python main.py --input yourfile.pdf")
        print("3. Or test with sample: python main.py --test")
        return 0
    else:
        print("\n Some tests failed. Please check the output above.")
        return 1

if __name__ == '__main__':
    sys.exit(main())
